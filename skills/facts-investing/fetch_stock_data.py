#!/usr/bin/env python3
"""
F.A.C.T.S Investment System - Stock Data Fetcher
自动获取股票数据并更新Excel表格
"""

import yfinance as yf
import openpyxl
import pandas as pd
from datetime import datetime
import sys
import os

# 配置
EXCEL_PATH = os.path.join(os.path.dirname(__file__), "BOS_Spreadsheet.xlsx")
BACKUP_PATH = os.path.join(os.path.dirname(__file__), f"backup/BOS_Spreadsheet_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")

def get_stock_data(ticker: str) -> dict:
    """获取单只股票的完整数据"""
    try:
        stock = yf.Ticker(ticker)
        info = stock.info
        
        # 获取财务数据
        financials = stock.financials
        balance_sheet = stock.balance_sheet
        cashflow = stock.cashflow
        
        # 基本信息
        current_price = info.get('currentPrice', info.get('regularMarketPrice', 0))
        pe_ratio = info.get('trailingPE', info.get('forwardPE', 0))
        
        # EPS (取最近4个季度)
        eps_trailing = info.get('trailingEps', 0)
        eps_forward = info.get('forwardEps', 0)
        
        # 收入和净利
        total_revenue = info.get('totalRevenue', 0)
        net_income = info.get('netIncomeToCommon', 0)
        
        # ROE
        roe = info.get('returnOnEquity', 0)
        if roe:
            roe = roe * 100  # 转为百分比
        
        # 毛利率和净利率
        gross_margin = info.get('grossMargins', 0)
        if gross_margin:
            gross_margin = gross_margin * 100
            
        profit_margin = info.get('profitMargins', 0)
        if profit_margin:
            profit_margin = profit_margin * 100
        
        # Debt
        total_debt = info.get('totalDebt', 0)
        total_cash = info.get('totalCash', 0)
        
        # Book Value
        book_value = info.get('bookValue', 0)
        
        # 市值
        market_cap = info.get('marketCap', 0)
        
        # 股息
        dividend_yield = info.get('dividendYield', 0)
        if dividend_yield:
            dividend_yield = dividend_yield * 100
        dividend_rate = info.get('dividendRate', 0)
        
        # 52周高低
        week_52_high = info.get('fiftyTwoWeekHigh', 0)
        week_52_low = info.get('fiftyTwoWeekLow', 0)
        
        # Beta
        beta = info.get('beta', 0)
        
        # Free Cash Flow (从cashflow表获取)
        try:
            fcf = cashflow.loc['Free Cash Flow'].iloc[0] if 'Free Cash Flow' in cashflow.index else 0
        except:
            fcf = 0
        
        # Interest Coverage (EBIT / Interest Expense)
        try:
            ebit = financials.loc['EBIT'].iloc[0] if 'EBIT' in financials.index else 0
            interest_expense = abs(financials.loc['Interest Expense'].iloc[0]) if 'Interest Expense' in financials.index else 0
            interest_coverage = ebit / interest_expense if interest_expense > 0 else float('inf')
        except:
            interest_coverage = 0
        
        # EPS 5年增长率
        try:
            eps_history = []
            for i in range(min(5, len(financials.columns))):
                if 'Diluted EPS' in financials.index:
                    eps_history.append(financials.loc['Diluted EPS'].iloc[i])
            if len(eps_history) >= 2 and eps_history[-1] != 0:
                eps_growth = ((eps_history[0] / eps_history[-1]) ** (1/len(eps_history)) - 1) * 100
            else:
                eps_growth = 0
        except:
            eps_growth = 0
        
        # Payout Ratio
        payout_ratio = info.get('payoutRatio', 0)
        if payout_ratio:
            payout_ratio = payout_ratio * 100
        
        return {
            'ticker': ticker,
            'current_price': current_price,
            'pe_ratio': pe_ratio,
            'eps_trailing': eps_trailing,
            'eps_forward': eps_forward,
            'total_revenue': total_revenue,
            'net_income': net_income,
            'roe': roe,
            'gross_margin': gross_margin,
            'profit_margin': profit_margin,
            'total_debt': total_debt,
            'total_cash': total_cash,
            'book_value': book_value,
            'market_cap': market_cap,
            'dividend_yield': dividend_yield,
            'dividend_rate': dividend_rate,
            'week_52_high': week_52_high,
            'week_52_low': week_52_low,
            'beta': beta,
            'fcf': fcf,
            'interest_coverage': interest_coverage,
            'eps_growth': eps_growth,
            'payout_ratio': payout_ratio,
            'fetch_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
    except Exception as e:
        print(f"Error fetching {ticker}: {e}")
        return None


def update_excel(stock_data: dict, row: int):
    """更新Excel中指定行的数据"""
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb['AC - Assess, Calculate']
        
        # Column D: Current Price
        ws.cell(row=row, column=4).value = stock_data['current_price']
        
        # Column Q: EPS (trailing)
        ws.cell(row=row, column=17).value = stock_data['eps_trailing']
        
        # 可以根据需要添加更多列...
        
        wb.save(EXCEL_PATH)
        wb.close()
        return True
    except Exception as e:
        print(f"Error updating Excel: {e}")
        return False


def generate_facts_analysis(stock_data: dict) -> dict:
    """根据F.A.C.T.S框架生成分析"""
    
    data = stock_data
    
    # Step 2: 10-Point Assessment
    assessment = {}
    
    # E - EPS consistency (简化: 检查5年)
    assessment['eps'] = 1 if data['eps_growth'] > 0 else 0
    
    # F - FCF positive
    assessment['fcf'] = 1 if data['fcf'] > 0 else 0
    
    # R - ROE > 15%
    assessment['roe'] = 1 if data['roe'] > 15 else 0
    
    # I - Interest Coverage
    if data['interest_coverage'] > 10:
        assessment['interest_coverage'] = 1
    elif data['interest_coverage'] > 4:
        assessment['interest_coverage'] = 0.5
    else:
        assessment['interest_coverage'] = 0
    
    # N - Net Margin > 10%
    if data['profit_margin'] > 20:
        assessment['net_margin'] = 1
    elif data['profit_margin'] > 10:
        assessment['net_margin'] = 0.5
    else:
        assessment['net_margin'] = 0
    
    # D - Dividends
    assessment['dividends'] = 1 if data['dividend_rate'] > 0 else 0
    
    # 计算总分
    base_score = sum(assessment.values())
    
    # Step 3: 估值
    pe = data['pe_ratio']
    eps = data['eps_trailing']
    growth = data['eps_growth']
    
    if pe and eps and growth and growth > 0:
        peg = pe / growth
        target_price = eps * (growth / 100)
        entry_price = target_price * 0.8
        exit_price = target_price * 1.2
    else:
        peg = None
        target_price = None
        entry_price = None
        exit_price = None
    
    return {
        'assessment': assessment,
        'base_score': base_score,
        'peg': peg,
        'target_price': target_price,
        'entry_price': entry_price,
        'exit_price': exit_price
    }


def print_analysis(stock_data: dict, analysis: dict):
    """打印分析结果"""
    
    data = stock_data
    
    print(f"\n{'='*50}")
    print(f"📊 {data['ticker']} F.A.C.T.S Analysis")
    print(f"{'='*50}")
    print(f"⏰ Data Fetched: {data['fetch_time']}")
    print()
    
    # 基本信息
    print(f"📈 Basic Info:")
    print(f"   Price: ${data['current_price']:.2f}")
    print(f"   PE: {data['pe_ratio']:.2f}" if data['pe_ratio'] else "   PE: N/A")
    print(f"   EPS: ${data['eps_trailing']:.2f}" if data['eps_trailing'] else "   EPS: N/A")
    print(f"   ROE: {data['roe']:.1f}%" if data['roe'] else "   ROE: N/A")
    print(f"   Net Margin: {data['profit_margin']:.1f}%" if data['profit_margin'] else "   Net Margin: N/A")
    print(f"   Dividend Yield: {data['dividend_yield']:.2f}%" if data['dividend_yield'] else "   Dividend Yield: 0%")
    print()
    
    # 10-Point Assessment
    print(f"✅ 10-Point Assessment:")
    assess = analysis['assessment']
    print(f"   E - EPS Growth:    {'✅' if assess['eps'] else '❌'} (+{assess['eps']})")
    print(f"   F - FCF Positive:  {'✅' if assess['fcf'] else '❌'} (+{assess['fcf']})")
    print(f"   R - ROE >15%:      {'✅' if assess['roe'] else '❌'} (+{assess['roe']})")
    print(f"   I - Interest Cov:  {'✅' if assess['interest_coverage'] else '❌'} (+{assess['interest_coverage']})")
    print(f"   N - Net Margin:    {'✅' if assess['net_margin'] else '❌'} (+{assess['net_margin']})")
    print(f"   D - Dividends:     {'✅' if assess['dividends'] else '❌'} (+{assess['dividends']})")
    print(f"   {'─'*40}")
    print(f"   Base Score: {analysis['base_score']}/6")
    print()
    
    # 估值
    print(f"💰 Valuation:")
    if analysis['peg'] is not None:
        print(f"   PEG: {analysis['peg']:.2f} {'✅ < 1' if analysis['peg'] < 1 else '⚠️ > 1'}")
        print(f"   Target Price: ${analysis['target_price']:.2f}")
        print(f"   Entry Price:  ${analysis['entry_price']:.2f} (20% discount)")
        print(f"   Exit Price:   ${analysis['exit_price']:.2f} (20% premium)")
    else:
        print(f"   ⚠️ Insufficient data for valuation")
    print()


def main():
    """主函数"""
    
    print("🦐 F.A.C.T.S Stock Data Fetcher")
    print("=" * 40)
    
    # 从Excel读取股票列表
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
        ws = wb['AC - Assess, Calculate']
        
        tickers = []
        for row in range(3, 25):
            ticker = ws.cell(row=row, column=1).value
            if ticker:
                tickers.append((ticker, row))
        
        wb.close()
        print(f"\n📋 Found {len(tickers)} stocks in portfolio")
        
    except Exception as e:
        print(f"Error reading Excel: {e}")
        # 使用默认股票列表
        tickers = [
            ('NVDA', 16), ('MSFT', 13), ('AAPL', 14), ('GOOG', 17),
            ('AMZN', 5), ('META', 9), ('NFLX', 10), ('BAC', 3),
            ('KO', 19), ('DIS', 12), ('AVGO', 4), ('PEP', 6),
            ('BABA', 11), ('ADBE', 18), ('QCOM', 15), ('NECB', 8),
            ('PARA', 7)
        ]
    
    # 获取命令行参数
    if len(sys.argv) > 1:
        # 只获取指定的股票
        requested = [t.upper() for t in sys.argv[1:]]
        tickers = [(t, row) for t, row in tickers if t in requested]
    
    if not tickers:
        print("⚠️ No stocks to fetch. Usage: python fetch_stock_data.py [TICKER]")
        return
    
    # 备份原文件
    os.makedirs(os.path.dirname(BACKUP_PATH), exist_ok=True)
    import shutil
    shutil.copy2(EXCEL_PATH, BACKUP_PATH)
    print(f"💾 Backup saved: {BACKUP_PATH}")
    
    # 获取数据
    results = []
    for ticker, row in tickers:
        print(f"\n🔄 Fetching {ticker}...")
        stock_data = get_stock_data(ticker)
        
        if stock_data:
            analysis = generate_facts_analysis(stock_data)
            print_analysis(stock_data, analysis)
            
            # 更新Excel
            if update_excel(stock_data, row):
                print(f"✅ Excel updated for {ticker}")
            
            results.append({
                'ticker': ticker,
                'data': stock_data,
                'analysis': analysis
            })
        else:
            print(f"❌ Failed to fetch {ticker}")
    
    # 汇总
    print(f"\n{'='*50}")
    print(f"📊 Summary: Updated {len(results)}/{len(tickers)} stocks")
    print(f"{'='*50}")
    
    # 生成CSV报告
    if results:
        report_path = os.path.join(os.path.dirname(__file__), f"reports/report_{datetime.now().strftime('%Y%m%d')}.csv")
        os.makedirs(os.path.dirname(report_path), exist_ok=True)
        
        df = pd.DataFrame([{
            'Ticker': r['ticker'],
            'Price': r['data']['current_price'],
            'PE': r['data']['pe_ratio'],
            'EPS': r['data']['eps_trailing'],
            'ROE': r['data']['roe'],
            'Net Margin': r['data']['profit_margin'],
            'Div Yield': r['data']['dividend_yield'],
            'Score': r['analysis']['base_score'],
            'PEG': r['analysis']['peg'],
            'Entry': r['analysis']['entry_price'],
            'Exit': r['analysis']['exit_price'],
            'Updated': r['data']['fetch_time']
        } for r in results])
        
        df.to_csv(report_path, index=False)
        print(f"\n📄 Report saved: {report_path}")
    
    print(f"\n✅ Done! Updated {EXCEL_PATH}")


if __name__ == "__main__":
    main()
